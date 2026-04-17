import streamlit as st
import io
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from auth import require_login
from supabase_client import carica_submissions, carica_registro, salva_registro, elimina_registro
from style import inject_css

require_login()
inject_css()

# ---------------------------------------------------------------------------
# Matrice Rischio Residuo
# ---------------------------------------------------------------------------
MATRICE = {
    ("ALTO",       "INADEGUATE"):               "ALTO",
    ("ALTO",       "PARZIALMENTE ADEGUATE"):    "ALTO",
    ("ALTO",       "PREVALENTEMENTE ADEGUATE"): "MEDIO ALTO",
    ("ALTO",       "ADEGUATE"):                 "MEDIO BASSO",
    ("MEDIO ALTO", "INADEGUATE"):               "ALTO",
    ("MEDIO ALTO", "PARZIALMENTE ADEGUATE"):    "MEDIO ALTO",
    ("MEDIO ALTO", "PREVALENTEMENTE ADEGUATE"): "MEDIO BASSO",
    ("MEDIO ALTO", "ADEGUATE"):                 "BASSO",
    ("MEDIO BASSO","INADEGUATE"):               "MEDIO ALTO",
    ("MEDIO BASSO","PARZIALMENTE ADEGUATE"):    "MEDIO BASSO",
    ("MEDIO BASSO","PREVALENTEMENTE ADEGUATE"): "MEDIO BASSO",
    ("MEDIO BASSO","ADEGUATE"):                 "BASSO",
    ("BASSO",      "INADEGUATE"):               "MEDIO BASSO",
    ("BASSO",      "PARZIALMENTE ADEGUATE"):    "BASSO",
    ("BASSO",      "PREVALENTEMENTE ADEGUATE"): "BASSO",
    ("BASSO",      "ADEGUATE"):                 "BASSO",
}

COLORI_RISCHIO = {
    "ALTO":        "#f44336",
    "MEDIO ALTO":  "#ff9800",
    "MEDIO BASSO": "#ffeb3b",
    "BASSO":       "#4caf50",
}

COLORI_EXCEL = {
    "ALTO":        "F44336",
    "MEDIO ALTO":  "FF9800",
    "MEDIO BASSO": "FFEB3B",
    "BASSO":       "4CAF50",
}


def calcola_residuo(inerente: str, misure: str) -> str:
    return MATRICE.get((inerente, misure), "N/D")


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------
def genera_excel_registro(righe: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro Rischi"

    intestazioni = [
        "ID", "Data", "Servizio TIC", "Fornitore TIC",
        "Funzione Essenziale", "Rischio Inerente",
        "Misure di Sicurezza", "Rischio Residuo",
    ]
    ws.append(intestazioni)
    for col_idx, _ in enumerate(intestazioni, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1976D2")
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(righe, 2):
        data = r.get("created_at", "")[:10] if r.get("created_at") else ""
        residuo = r.get("rischio_residuo", "")
        ws.append([
            str(i - 1),
            data,
            r.get("servizio_tic", ""),
            r.get("fornitore_tic", ""),
            "Sì" if r.get("funzione_essenziale") else "No",
            r.get("rischio_inerente", ""),
            r.get("misure_sicurezza", ""),
            residuo,
        ])
        colore = COLORI_EXCEL.get(residuo)
        if colore:
            ws.cell(row=i, column=8).fill = PatternFill("solid", fgColor=colore)
            ws.cell(row=i, column=8).font = Font(bold=True)

    for col_idx in range(1, len(intestazioni) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# APP
# ===========================================================================
st.title("Registro Rischi ICT")
st.markdown("Riepilogo di tutti i servizi TIC valutati con il relativo rischio residuo.")
st.divider()

# ---------------------------------------------------------------------------
# Carica dati
# ---------------------------------------------------------------------------
registro = carica_registro()
q1_list = carica_submissions(tipo=1)
q2_list = carica_submissions(tipo=2)

# ---------------------------------------------------------------------------
# Tabella riepilogativa
# ---------------------------------------------------------------------------
if registro:
    st.subheader(f"Registro ({len(registro)} voci)")

    # Export Excel
    xlsx = genera_excel_registro(registro)
    st.download_button(
        "Scarica Excel",
        data=xlsx,
        file_name=f"registro_rischi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("")

    # Intestazioni tabella
    hdr = st.columns([0.5, 1.2, 2, 2, 1, 1.3, 1.8, 1.3, 0.6])
    for col, label in zip(hdr, ["#", "Data", "Servizio TIC", "Fornitore TIC",
                                  "F. Ess.", "Rischio Inerente",
                                  "Misure Sicurezza", "Rischio Residuo", ""]):
        col.markdown(f"**{label}**")

    for idx, r in enumerate(registro, 1):
        residuo = r.get("rischio_residuo", "N/D")
        colore = COLORI_RISCHIO.get(residuo, "#9e9e9e")
        data = r.get("created_at", "")[:10] if r.get("created_at") else ""

        cols = st.columns([0.5, 1.2, 2, 2, 1, 1.3, 1.8, 1.3, 0.6])
        cols[0].write(str(idx))
        cols[1].write(data)
        cols[2].write(r.get("servizio_tic", ""))
        cols[3].write(r.get("fornitore_tic", ""))
        cols[4].write("Sì" if r.get("funzione_essenziale") else "No")
        cols[5].write(r.get("rischio_inerente", ""))
        cols[6].write(r.get("misure_sicurezza", ""))
        cols[7].markdown(
            f'<span style="background-color:{colore}; color:{"black" if residuo == "MEDIO BASSO" else "white"}; '
            f'padding:3px 8px; border-radius:6px; font-weight:bold; font-size:13px;">'
            f'{residuo}</span>',
            unsafe_allow_html=True,
        )
        if cols[8].button("🗑", key=f"del_{r['id']}"):
            elimina_registro(r["id"])
            st.rerun()

else:
    st.info("Nessuna voce nel registro. Aggiungi la prima voce qui sotto.")

st.divider()

# ---------------------------------------------------------------------------
# Form aggiunta nuova voce
# ---------------------------------------------------------------------------
st.subheader("Aggiungi nuova voce")

if not q1_list:
    st.warning("Nessun Questionario 1 compilato. Compila prima il Questionario 1.")
elif not q2_list:
    st.warning("Nessun Questionario 2 compilato. Compila prima il Questionario 2.")
else:
    with st.form("form_registro"):
        col1, col2 = st.columns(2)

        # Selezione Q1
        q1_opzioni = {
            f"{r['nome_servizio_ict']} — {r['livello_rischio']} ({r['created_at'][:10]})": r
            for r in q1_list
        }
        with col1:
            q1_sel_label = st.selectbox("Questionario 1 (Rischio ICT)", options=list(q1_opzioni.keys()))

        # Selezione Q2
        q2_opzioni = {
            f"{r['nome_servizio_ict']} — {r['livello_rischio']} ({r['created_at'][:10]})": r
            for r in q2_list
        }
        with col2:
            q2_sel_label = st.selectbox("Questionario 2 (Sicurezza Fornitore)", options=list(q2_opzioni.keys()))

        q1_sel = q1_opzioni[q1_sel_label]
        q2_sel = q2_opzioni[q2_sel_label]

        # Campi manuali
        col3, col4 = st.columns(2)
        with col3:
            servizio_tic = st.text_input("Nome Servizio TIC", value=q1_sel["nome_servizio_ict"])
        with col4:
            fornitore_tic = st.text_input("Nome Fornitore TIC", value=q2_sel["nome_servizio_ict"])

        funzione_ess = st.radio(
            "Il servizio è a supporto di una funzione essenziale/importante?",
            options=["Sì", "No"],
            horizontal=True,
        ) == "Sì"

        # Anteprima calcolo
        rischio_inerente = q1_sel["livello_rischio"]
        misure_sicurezza = q2_sel["livello_rischio"]
        rischio_residuo = calcola_residuo(rischio_inerente, misure_sicurezza)
        colore_prev = COLORI_RISCHIO.get(rischio_residuo, "#9e9e9e")

        st.markdown("**Anteprima calcolo:**")
        c1, c2, c3 = st.columns(3)
        c1.metric("Rischio Inerente (Q1)", rischio_inerente)
        c2.metric("Misure Sicurezza (Q2)", misure_sicurezza)
        c3.markdown(
            f'<div style="text-align:center; padding:10px; border-radius:8px; '
            f'background-color:{colore_prev}; color:{"black" if rischio_residuo == "MEDIO BASSO" else "white"}; '
            f'font-weight:bold; font-size:16px; margin-top:8px;">Rischio Residuo<br>{rischio_residuo}</div>',
            unsafe_allow_html=True,
        )

        submitted = st.form_submit_button("Aggiungi al Registro", type="primary", use_container_width=True)

    if submitted:
        ok = salva_registro(
            servizio_tic=servizio_tic,
            fornitore_tic=fornitore_tic,
            funzione_essenziale=funzione_ess,
            submission_q1_id=q1_sel["id"],
            submission_q2_id=q2_sel["id"],
            rischio_inerente=rischio_inerente,
            misure_sicurezza=misure_sicurezza,
            rischio_residuo=rischio_residuo,
        )
        if ok:
            st.success("Voce aggiunta al registro.")
            st.rerun()

# Logout sidebar
with st.sidebar:
    st.markdown("---")
    if st.button("Esci", use_container_width=True):
        st.session_state["logged_in"] = False
        st.rerun()
