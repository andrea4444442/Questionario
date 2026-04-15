import streamlit as st
import io
import os
import sys
from datetime import date, datetime
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from auth import require_login

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Logo Meta.png")

require_login()


@st.cache_data(ttl=0)
def carica_logo_bytes(mtime: float) -> bytes | None:
    if not os.path.exists(LOGO_PATH):
        return None
    with open(LOGO_PATH, "rb") as f:
        return f.read()


def _logo_bytes() -> bytes | None:
    if not os.path.exists(LOGO_PATH):
        return None
    mtime = os.path.getmtime(LOGO_PATH)
    return carica_logo_bytes(mtime)


# ---------------------------------------------------------------------------
# Generazione PDF
# ---------------------------------------------------------------------------

def _san(text: str) -> str:
    """Sostituisce caratteri Unicode non supportati da Helvetica (Latin-1)."""
    return (
        str(text)
        .replace("\u2019", "'").replace("\u2018", "'")
        .replace("\u201c", '"').replace("\u201d", '"')
        .replace("\u2013", "-").replace("\u2014", "-")
        .replace("\u2026", "...").replace("\u2192", "->")
        .replace("\u2190", "<-").replace("\u00b7", "-")
    )


class PDFReport(FPDF):
    def header(self):
        y_start = self.y
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, "Questionario Sicurezza ICT Fornitori", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 6, f"Data: {date.today().strftime('%d/%m/%Y')}", new_x="LMARGIN", new_y="NEXT")
        logo_bytes = _logo_bytes()
        if logo_bytes:
            self.image(io.BytesIO(logo_bytes), x=self.w - 60, y=y_start, w=50)
        self.set_x(self.l_margin)
        self.ln(6)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}", align="C")


def _pdf_sezione(pdf, titolo: str, voci: list[tuple[str, str]]) -> None:
    w = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 8, _san(titolo))
    pdf.set_font("Helvetica", "", 9)
    for label, valore in voci:
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(w, 5, _san(f"  {label}"))
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(w, 5, _san(f"  -> {valore}"))
        pdf.set_font("Helvetica", "", 9)
        pdf.ln(1)
    pdf.ln(3)


def genera_pdf(data: dict) -> bytes:
    pdf = PDFReport()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    w = pdf.w - pdf.l_margin - pdf.r_margin

    _pdf_sezione(pdf, "Informazioni Generali", [
        ("Ragione Sociale", data["ragione_sociale"]),
        ("Descrizione servizi/prodotti", data["descrizione"]),
        ("Paese fornitura", data["paese"]),
        ("Funzione essenziale/importante", "Vero" if data["funzione_essenziale"] else "Falso"),
    ])

    for sez in data["sezioni"]:
        _pdf_sezione(pdf, sez["titolo"], sez["voci"])
        if sez.get("punteggio") is not None:
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(w, 6, _san(f"  Punteggio sezione: {sez['punteggio']}"))
            pdf.ln(2)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 10, _san(f"PUNTEGGIO TOTALE: {data['punteggio_totale']}"))
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 10, _san(f"MISURE DI SICUREZZA: {data['valutazione']} (valore: {data['valore']})"))
    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Generazione Excel
# ---------------------------------------------------------------------------

def genera_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionario"

    ws.append(["Campo", "Valore"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.append(["Ragione Sociale", data["ragione_sociale"]])
    ws.append(["Descrizione", data["descrizione"]])
    ws.append(["Paese", data["paese"]])
    ws.append(["Funzione essenziale", "Vero" if data["funzione_essenziale"] else "Falso"])
    ws.append([])

    for sez in data["sezioni"]:
        ws.append([sez["titolo"]])
        ws[f"A{ws.max_row}"].font = Font(bold=True)
        for label, valore in sez["voci"]:
            ws.append([label, valore])
        if sez.get("punteggio") is not None:
            ws.append(["Punteggio sezione", sez["punteggio"]])
        ws.append([])

    ws.append(["PUNTEGGIO TOTALE", data["punteggio_totale"]])
    ws.append(["VALUTAZIONE MISURE DI SICUREZZA", data["valutazione"]])
    ws.append(["VALORE", data["valore"]])
    last = ws.max_row
    for row in range(last - 2, last + 1):
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws[f"B{row}"].font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helper: upload + download inline
# ---------------------------------------------------------------------------

def file_upload_con_download(label: str, key: str, tipi: list[str] | None = None) -> str:
    """Mostra un file uploader, salva i bytes in session_state, restituisce il nome file."""
    f = st.file_uploader(label, key=key, type=tipi)
    sk_bytes = f"_file_bytes_{key}"
    sk_name = f"_file_name_{key}"
    if f is not None:
        st.session_state[sk_bytes] = f.getvalue()
        st.session_state[sk_name] = f.name
    if sk_bytes in st.session_state:
        st.download_button(
            f"⬇ Scarica: {st.session_state[sk_name]}",
            data=st.session_state[sk_bytes],
            file_name=st.session_state[sk_name],
            key=f"dl_{key}",
        )
        return st.session_state[sk_name]
    return "Non allegato"


# ===========================================================================
# APP PRINCIPALE
# ===========================================================================

st.title("Questionario Sicurezza ICT Fornitori")
st.markdown("Valutazione delle misure di sicurezza ICT del fornitore.")
st.divider()

# ---------------------------------------------------------------------------
# 1. Informazioni Generali
# ---------------------------------------------------------------------------
st.header("1. Informazioni Generali")

ragione_sociale = st.text_input("Ragione Sociale")
descrizione = st.text_area("Descrizione servizi/prodotti")
paese = st.text_input("Paese in cui ha luogo la fornitura del/i prodotto/servizi")
funzione_essenziale = (
    st.radio(
        "Il servizio TIC fornito è a supporto di una funzione essenziale o importante?",
        options=["Vero", "Falso"],
        index=0,
        horizontal=True,
        key="q2_funzione",
    )
    == "Vero"
)

st.divider()

# Accumulatori
punti_totali = 0
sezioni_data: list[dict] = []

# ---------------------------------------------------------------------------
# 2. Protezione dei Dati  (comune a VERO e FALSO)
# ---------------------------------------------------------------------------
st.header("2. Protezione dei Dati")

trattamento_ue = st.radio(
    "Trattamento dati in UE",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_ue",
)
p_prot = 1 if trattamento_ue == "Sì" else 0
punti_totali += p_prot
voci_prot: list[tuple[str, str]] = [("Trattamento dati in UE", trattamento_ue)]

if trattamento_ue == "Sì":
    luogo_gestione = st.text_input("Luogo di gestione dei dati (trattamento)", key="q2_luogo")
    ubicazione_riposo = st.text_input("Ubicazione dei dati a riposo (archiviazione)", key="q2_ubicazione")
    durata = st.number_input("Durata del trattamento (in giorni)", min_value=0, step=1, key="q2_durata")
    tratt_personali = st.text_input("Il servizio prevede il trattamento di Dati Personali?", key="q2_personali")
    natura_scopo = st.text_area("Natura e scopo del trattamento", key="q2_natura")
    cat_soggetti = st.text_input("Categorie dei soggetti interessati", key="q2_cat_soggetti")
    cat_dati = st.text_input("Categorie di dati personali", key="q2_cat_dati")
    nome_dpia = file_upload_con_download(
        "Allegare il DPIA per il trattamento di dati sensibili",
        key="q2_dpia",
        tipi=["pdf", "docx", "xlsx", "png", "jpg"],
    )
    voci_prot += [
        ("Luogo di gestione dei dati", luogo_gestione),
        ("Ubicazione dati a riposo", ubicazione_riposo),
        ("Durata trattamento (giorni)", str(int(durata))),
        ("Trattamento Dati Personali", tratt_personali),
        ("Natura e scopo del trattamento", natura_scopo),
        ("Categorie soggetti interessati", cat_soggetti),
        ("Categorie dati personali", cat_dati),
        ("DPIA allegato", nome_dpia),
    ]

sezioni_data.append({"titolo": "Protezione dei Dati", "voci": voci_prot, "punteggio": p_prot})
st.divider()

# ===========================================================================
# RAMO VERO
# ===========================================================================
if funzione_essenziale:

    # -----------------------------------------------------------------------
    # 3. Certificazioni e Standard
    # -----------------------------------------------------------------------
    st.header("3. Certificazioni e Standard")
    p_cert = 0
    voci_cert: list[tuple[str, str]] = []

    isms = st.radio(
        "Avete un sistema di gestione della sicurezza informatica (ISMS) o simile?",
        options=["Sì", "No"],
        index=0,
        horizontal=True,
        key="q2_isms",
    )
    voci_cert.append(("ISMS o sistema equivalente", isms))

    if isms == "Sì":
        CERT_PUNTI = {
            "ISO 27701": 2,
            "ISO 27001": 1,
            "SOC 2 (ISAE 3000)": 3,
            "BSI CS": 3,
            "Altro": 1,
            "Nessuna": 0,
        }
        cert_sel = st.multiselect(
            "Quali delle seguenti certificazioni possedete attualmente in materia di sicurezza informatica? (selezione multipla)",
            options=list(CERT_PUNTI.keys()),
            key="q2_cert",
        )
        for c in cert_sel:
            if c != "Nessuna":
                p_cert += CERT_PUNTI[c]
        voci_cert.append(("Certificazioni possedute", ", ".join(cert_sel) if cert_sel else "Nessuna selezionata"))

        if "Altro" in cert_sel:
            altre = st.text_input(
                "Si prega di identificare le certificazioni possedute dalla vostra azienda",
                key="q2_altre_cert",
            )
            voci_cert.append(("Dettaglio certificazioni 'Altro'", altre))

    # Sempre chiesto
    revisione = st.radio(
        "Queste certificazioni vengono revisionate a cadenza annuale da un ente esterno?",
        options=["Sì", "No"],
        index=0,
        horizontal=True,
        key="q2_revisione",
    )
    if revisione == "Sì":
        p_cert += 1
    voci_cert.append(("Revisione annuale da ente esterno", revisione))

    nome_audit = file_upload_con_download(
        "Si prega di fornire i risultati dell'ultimo Audit (anche interno)",
        key="q2_audit",
        tipi=["pdf", "docx", "xlsx"],
    )
    voci_cert.append(("Risultati ultimo Audit allegati", nome_audit))

    punti_totali += p_cert
    sezioni_data.append({"titolo": "Certificazioni e Standard", "voci": voci_cert, "punteggio": p_cert})
    st.divider()

    # -----------------------------------------------------------------------
    # 4. Governance e Gestione Rischi ICT
    # -----------------------------------------------------------------------
    st.header("4. Governance e Gestione Rischi ICT")
    p_gov = 0
    voci_gov: list[tuple[str, str]] = []

    policy_doc = st.radio(
        "Disponete di Politiche di Sicurezza Informatica documentate?",
        options=["Sì", "No"],
        index=0,
        horizontal=True,
        key="q2_policy_doc",
    )
    if policy_doc == "Sì":
        p_gov += 1
    voci_gov.append(("Politiche di Sicurezza documentate", policy_doc))

    nome_policy = file_upload_con_download(
        "Caricate la Policy di sicurezza informatica o un documento con i controlli di sicurezza",
        key="q2_policy_file",
        tipi=["pdf", "docx", "xlsx"],
    )
    voci_gov.append(("Policy sicurezza allegata", nome_policy))

    formazione = st.radio(
        "Queste politiche prevedono una formazione periodica del personale sulla cybersecurity?",
        options=["Sì", "No"],
        index=0,
        horizontal=True,
        key="q2_formazione",
    )
    if formazione == "Sì":
        p_gov += 1
    voci_gov.append(("Formazione periodica cybersecurity", formazione))

    freq_audit_map = {"Trimestrale": 2, "Annuale": 1, "Minore di annuale": 0}
    freq_audit = st.selectbox(
        "Con quale frequenza eseguite audit di sicurezza interni o esterni?",
        options=list(freq_audit_map.keys()),
        key="q2_freq_audit",
    )
    p_gov += freq_audit_map[freq_audit]
    voci_gov.append(("Frequenza audit di sicurezza", freq_audit))

    punti_totali += p_gov
    sezioni_data.append({"titolo": "Governance e Gestione Rischi ICT", "voci": voci_gov, "punteggio": p_gov})
    st.divider()

    # -----------------------------------------------------------------------
    # 5. Gestione Fornitori e Sicurezza Supply Chain
    # -----------------------------------------------------------------------
    st.header("5. Gestione Fornitori e Sicurezza Supply Chain")
    p_forn = 0
    voci_forn: list[tuple[str, str]] = []

    conformita = st.radio(
        "Monitorate la conformità dei fornitori terzi ed eventuali sub-fornitori alle vostre politiche di sicurezza?",
        options=["Sì", "No"],
        index=0,
        horizontal=True,
        key="q2_conformita",
    )
    if conformita == "Sì":
        p_forn += 1
    voci_forn.append(("Monitoraggio conformità fornitori", conformita))

    RISCHI_TERZI = {
        "Nel contratto": 1,
        "Monitoraggio dei livelli di servizio e dei controlli di sicurezza": 1,
        "Valutazioni della sicurezza annuali o trimestrali": 1,
        "Altro": 0,
        "Nessuno": 0,
    }
    rischi_terzi = st.multiselect(
        "In che modo gestite i rischi di terzi? (selezione multipla)",
        options=list(RISCHI_TERZI.keys()),
        key="q2_rischi_terzi",
    )
    for r in rischi_terzi:
        p_forn += RISCHI_TERZI[r]
    voci_forn.append(("Gestione rischi di terzi", ", ".join(rischi_terzi) if rischi_terzi else "Non specificato"))

    if "Altro" in rischi_terzi:
        desc_rischi = st.text_area(
            "Si prega di descrivere le misure per la gestione dei rischi di terzi, compresi il rischio di sub-esternalizzazione",
            key="q2_desc_rischi",
        )
        voci_forn.append(("Dettaglio misure rischi terzi", desc_rischi))

    punti_totali += p_forn
    sezioni_data.append({"titolo": "Gestione Fornitori e Supply Chain", "voci": voci_forn, "punteggio": p_forn})
    st.divider()

# ===========================================================================
# SEZIONI COMUNI (VERO e FALSO)
# ===========================================================================
_base_num = 6 if funzione_essenziale else 3

# -----------------------------------------------------------------------
# Continuità Operativa e Disaster Recovery
# -----------------------------------------------------------------------
st.header(f"{_base_num}. Continuità Operativa e Disaster Recovery")
p_cont = 0
voci_cont: list[tuple[str, str]] = []

piano_risposta = st.radio(
    "Disponete di un piano di risposta agli incidenti e di un Piano di Continuità Operativa (PCO)?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_piano_risposta",
)
if piano_risposta == "Sì":
    p_cont += 1
voci_cont.append(("Piano risposta incidenti e PCO", piano_risposta))

nome_pco = file_upload_con_download(
    "Caricare il piano di risposta agli incidenti e/o il PCO",
    key="q2_pco",
    tipi=["pdf", "docx", "xlsx"],
)
voci_cont.append(("PCO allegato", nome_pco))

backup = st.radio(
    "La Società è dotata di policy e procedure per il backup e restore dei dati?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_backup",
)
if backup == "Sì":
    p_cont += 1
voci_cont.append(("Policy backup e restore", backup))

punti_totali += p_cont
sezioni_data.append({"titolo": "Continuità Operativa e Disaster Recovery", "voci": voci_cont, "punteggio": p_cont})
st.divider()

# -----------------------------------------------------------------------
# Gestione e Risposta agli Incidenti ICT
# -----------------------------------------------------------------------
st.header(f"{_base_num + 1}. Gestione e Risposta agli Incidenti ICT")
p_inc = 0
voci_inc: list[tuple[str, str]] = []

pol_incidenti = st.radio(
    "Disponete di politiche per la gestione degli incidenti ICT e di sicurezza in linea con gli standard e le best practice di settore (es. ISO 27001)?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_pol_inc",
)
if pol_incidenti == "Sì":
    p_inc += 1
voci_inc.append(("Politiche gestione incidenti ICT", pol_incidenti))

comm_clienti = st.radio(
    "È prevista una comunicazione strutturata ai clienti in caso di incidenti che impattano sui servizi erogati, che preveda aggiornamenti sullo stato dell'incidente, sulle previsioni e sulle contromisure adottate?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_comm_clienti",
)
if comm_clienti == "Sì":
    p_inc += 1
voci_inc.append(("Comunicazione strutturata ai clienti", comm_clienti))

incidenti = st.radio(
    "Negli ultimi due anni sono stati rilevati incidenti operativi e di sicurezza rilevanti riferibili alla propria infrastruttura o a fornitori?",
    options=["Sì", "No"],
    index=1,
    horizontal=True,
    key="q2_incidenti",
)
if incidenti == "No":
    p_inc += 1
voci_inc.append(("Incidenti rilevati ultimi 2 anni", incidenti))

if incidenti == "Sì":
    desc_inc = st.text_area(
        "Si prega di descrivere la tipologia di incidenti che si è verificata",
        key="q2_desc_inc",
    )
    voci_inc.append(("Descrizione incidenti verificati", desc_inc))

punti_totali += p_inc
sezioni_data.append({"titolo": "Gestione e Risposta agli Incidenti ICT", "voci": voci_inc, "punteggio": p_inc})
st.divider()

# -----------------------------------------------------------------------
# Gestione degli Asset e Sicurezza Fisica
# -----------------------------------------------------------------------
st.header(f"{_base_num + 2}. Gestione degli Asset e Sicurezza Fisica")
p_asset = 0
voci_asset: list[tuple[str, str]] = []

presidi = st.radio(
    "Disponete di presidi per l'accesso alle proprie sedi consentendolo al solo personale autorizzato (almeno attraverso l'uso di badge personali)?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_presidi",
)
if presidi == "Sì":
    p_asset += 1
voci_asset.append(("Presidi accesso sedi (badge personali)", presidi))

protezione_dc = st.radio(
    "Garantite un adeguato livello di protezione fisica dei propri data center?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_dc",
)
if protezione_dc == "Sì":
    p_asset += 1
voci_asset.append(("Protezione fisica data center", protezione_dc))

dispositivi = st.radio(
    "Disponete di misure di sicurezza per la gestione dei dispositivi mobili e portatili (crittografia, backup etc.)?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_dispositivi",
)
if dispositivi == "Sì":
    p_asset += 1
voci_asset.append(("Sicurezza dispositivi mobili e portatili", dispositivi))

punti_totali += p_asset
sezioni_data.append({"titolo": "Gestione Asset e Sicurezza Fisica", "voci": voci_asset, "punteggio": p_asset})
st.divider()

# -----------------------------------------------------------------------
# Identity and Access Management
# -----------------------------------------------------------------------
st.header(f"{_base_num + 3}. Identity and Access Management")
p_iam = 0
voci_iam: list[tuple[str, str]] = []

governo_acc = st.radio(
    "Disponete di un sistema di governo degli accessi ai propri sistemi secondo i principi di 'need to know' e 'need to use' (segregation of duties), con processi sicuri per la creazione, gestione e revoca delle credenziali?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_gov_acc",
)
if governo_acc == "Sì":
    p_iam += 1
voci_iam.append(("Sistema governo accessi (segregation of duties)", governo_acc))

vpn_mfa = st.radio(
    "L'accesso dalla rete e da remoto è effettuato tramite VPN e sistemi di autenticazione multifattore?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_vpn_mfa",
)
if vpn_mfa == "Sì":
    p_iam += 1
voci_iam.append(("VPN e autenticazione multifattore", vpn_mfa))

endpoint = st.radio(
    "Sono previste soluzioni di protezione degli endpoint (antivirus/EDR) centralizzate con regolare aggiornamento delle firme antivirus (frequenza almeno giornaliera)?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_endpoint",
)
if endpoint == "Sì":
    p_iam += 1
voci_iam.append(("Protezione endpoint centralizzata (antivirus/EDR)", endpoint))

acc_logico = st.radio(
    "Disponete di misure di sicurezza adeguate alle best practice di settore (es. ISO27001, NIST) per l'accesso logico agli ambienti non produttivi e produttivi da parte di terze parti?",
    options=["Sì", "No"],
    index=0,
    horizontal=True,
    key="q2_acc_logico",
)
if acc_logico == "Sì":
    p_iam += 1
voci_iam.append(("Accesso logico ambienti da terze parti", acc_logico))

punti_totali += p_iam
sezioni_data.append({"titolo": "Identity and Access Management", "voci": voci_iam, "punteggio": p_iam})
st.divider()

# ===========================================================================
# RAMO VERO — Change Management e Vulnerability Management
# ===========================================================================
if funzione_essenziale:

    st.header(f"{_base_num + 4}. Change Management")
    p_chg = 0
    voci_chg: list[tuple[str, str]] = []

    change = st.radio(
        "Disponete di un processo di change management che garantisce al Cliente il controllo su modifiche, sostituzioni o adeguamenti tecnologici dei servizi erogati, seguendo procedure formalizzate, con lo scopo di ridurre i rischi e minimizzare l'impatto di possibili incidenti?",
        options=["Sì", "No"],
        index=0,
        horizontal=True,
        key="q2_change",
    )
    if change == "Sì":
        p_chg += 1
    voci_chg.append(("Processo di change management formalizzato", change))

    punti_totali += p_chg
    sezioni_data.append({"titolo": "Change Management", "voci": voci_chg, "punteggio": p_chg})
    st.divider()

    st.header(f"{_base_num + 5}. Vulnerability Management")
    p_vuln = 0
    voci_vuln: list[tuple[str, str]] = []

    vuln = st.radio(
        "Effettuate almeno annualmente vulnerability/penetration test su applicazioni e sistemi e predisponete ed implementate tempestivamente gli eventuali piani di trattamento? Vi impegnate a supportare il cliente nella gestione dei Threat Led Penetration Test (TLPT) ove richiesto?",
        options=["Sì", "No"],
        index=0,
        horizontal=True,
        key="q2_vuln",
    )
    if vuln == "Sì":
        p_vuln += 1
    voci_vuln.append(("Vulnerability/penetration test almeno annuali", vuln))

    punti_totali += p_vuln
    sezioni_data.append({"titolo": "Vulnerability Management", "voci": voci_vuln, "punteggio": p_vuln})
    st.divider()

# ===========================================================================
# RISULTATO
# ===========================================================================
_last_num = (_base_num + 6) if funzione_essenziale else (_base_num + 4)
st.header(f"{_last_num}. Risultato")

if st.button("Calcola Valutazione", type="primary", use_container_width=True):
    if not ragione_sociale.strip():
        st.warning("Inserire la Ragione Sociale prima di procedere.")
    else:
        if funzione_essenziale:
            if punti_totali >= 26:
                valutazione, valore, colore = "ADEGUATE", 4, "#4caf50"
            elif punti_totali >= 18:
                valutazione, valore, colore = "PREVALENTEMENTE ADEGUATE", 3, "#8bc34a"
            elif punti_totali >= 9:
                valutazione, valore, colore = "PARZIALMENTE ADEGUATE", 2, "#ff9800"
            else:
                valutazione, valore, colore = "INADEGUATE", 1, "#f44336"
        else:
            if punti_totali >= 11:
                valutazione, valore, colore = "ADEGUATE", 4, "#4caf50"
            elif punti_totali >= 7:
                valutazione, valore, colore = "PREVALENTEMENTE ADEGUATE", 3, "#8bc34a"
            elif punti_totali >= 4:
                valutazione, valore, colore = "PARZIALMENTE ADEGUATE", 2, "#ff9800"
            else:
                valutazione, valore, colore = "INADEGUATE", 1, "#f44336"

        st.session_state["risultato_q2"] = {
            "ragione_sociale": ragione_sociale,
            "descrizione": descrizione,
            "paese": paese,
            "funzione_essenziale": funzione_essenziale,
            "sezioni": sezioni_data,
            "punteggio_totale": punti_totali,
            "valutazione": valutazione,
            "valore": valore,
            "colore": colore,
            "export_id": datetime.now().strftime("%Y%m%d_%H%M%S"),
        }

if "risultato_q2" in st.session_state:
    res = st.session_state["risultato_q2"]

    st.markdown("---")
    st.subheader("Riepilogo Punteggi per Sezione")
    for sez in res["sezioni"]:
        if sez.get("punteggio") is not None:
            st.markdown(f"- **{sez['titolo']}**: {sez['punteggio']} punti")

    st.markdown(
        f'<div style="text-align:center; padding:20px; border-radius:10px; '
        f'background-color:{res["colore"]}; color:white; font-size:24px; font-weight:bold;">'
        f'MISURE DI SICUREZZA: {res["valutazione"]}<br>'
        f'<span style="font-size:16px;">'
        f'Punteggio: {res["punteggio_totale"]} &nbsp;|&nbsp; Valore: {res["valore"]}'
        f'</span></div>',
        unsafe_allow_html=True,
    )

    st.markdown("")
    export_id = res.get("export_id", datetime.now().strftime("%Y%m%d_%H%M%S"))
    nome_file = f"sicurezza_ict_{res['ragione_sociale'].replace(' ', '_')}_{export_id}"

    col_pdf, col_xlsx = st.columns(2)
    with col_pdf:
        pdf_bytes = genera_pdf(res)
        st.download_button(
            "Scarica PDF",
            data=pdf_bytes,
            file_name=f"{nome_file}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    with col_xlsx:
        xlsx_bytes = genera_excel(res)
        st.download_button(
            "Scarica Excel",
            data=xlsx_bytes,
            file_name=f"{nome_file}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
