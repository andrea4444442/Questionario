import streamlit as st
import io
import os
import sys
from datetime import date, datetime
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XlImage

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from auth import require_login
from supabase_client import salva_submission
from style import inject_css

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Logo Meta.png")

require_login()
inject_css()


@st.cache_data(ttl=0)
def carica_logo_bytes(mtime: float) -> bytes | None:
    """Carica il logo dal disco. Il parametro mtime forza il ricaricamento se il file cambia."""
    if not os.path.exists(LOGO_PATH):
        return None
    with open(LOGO_PATH, "rb") as f:
        return f.read()


def _logo_bytes() -> bytes | None:
    """Restituisce sempre i bytes aggiornati del logo corrente."""
    if not os.path.exists(LOGO_PATH):
        return None
    mtime = os.path.getmtime(LOGO_PATH)
    return carica_logo_bytes(mtime)

# ---------------------------------------------------------------------------
# Dati domande per scenario
# ---------------------------------------------------------------------------

DOMANDE_DISPONIBILITA = [
    "L'eventuale interruzione del servizio comporta penali contrattuali verso clienti, perdita di commissioni/interessi, incremento costi straordinari (soluzioni temporanee, interventi di emergenza, consulenze esterne), possibile sanzione da parte dell'Autorità di Vigilanza?",
    "L'eventuale interruzione del servizio comporta reclami e disservizi diffusi, esposizione negativa su media/social, percezione di inaffidabilità?",
    "L'eventuale interruzione del servizio comporta difficoltà a sostenere iniziative di digital banking, perdita di clienti a favore di concorrenti più resilienti, ostacoli nei rapporti con partner strategici?",
    "L'eventuale interruzione del servizio TIC comporta blocchi nelle operazioni di pagamento, impossibilità di accesso a home/mobile banking, malfunzionamento ATM e POS, ritardi nelle attività di tesoreria?",
]

DOMANDE_SICUREZZA = [
    "L'eventuale accesso non autorizzato potrebbe comportare il pagamento riscatti, risarcimenti ai clienti, multe da parte delle autorità di vigilanza, incremento costi di sicurezza?",
    "L'eventuale accesso non autorizzato potrebbe comportare reclami diffusi, pubblicità negativa su media/social, calo acquisizione di nuovi clienti, segnalazioni pubbliche del Garante Privacy?",
    "L'eventuale accesso non autorizzato/disfunzione del servizio potrebbe comportare difficoltà nel lanciare nuovi servizi digitali, perdita clienti verso concorrenti fintech, rischio di esclusione da partnership strategiche?",
    "L'eventuale degrado o perdita di integrità/sicurezza del servizio potrebbe comportare blocchi nei sistemi di pagamento, interruzione home banking, compromissione dei canali ATM/POS?",
]

DOMANDE_CAMBIAMENTI = [
    "L'eventuale interruzione del servizio TIC dovuto al processo di change ICT potrebbe comportare extra costi per ripristini d'urgenza, cause con fornitori/partner, riduzione dei ricavi per servizi non disponibili?",
    "L'eventuale interruzione del servizio TIC correlato al processo di change potrebbe comportare reclami massivi, articoli negativi sulla stampa, percezione di scarsa affidabilità tecnologica?",
    "L'eventuale interruzione del servizio TIC dovuto al processo di change ICT comporta ritardi nei programmi di digital banking, fallimento di iniziative fintech/innovative, abbandono da parte di partner tecnologici?",
    "Il servizio TIC è soggetto a frequenti aggiornamenti normativi e tecnologici?",
]

DOMANDE_INTEGRITA = [
    "L'eventuale perdita di integrità dei dati del servizio TIC può comportare errori nei saldi contabili, anomalie nei bonifici, rendicontazioni regolamentari non affidabili?",
    "L'eventuale perdita di integrità dei dati del servizio TIC può comportare rimborsi a clienti per transazioni errate, multe delle autorità di vigilanza, costi di ripristino e revisione massiva delle basi dati?",
    "L'eventuale perdita di integrità dei dati del servizio TIC può comportare reclami per errori di saldo, percezione di inefficienza o frode, esposizione mediatica negativa?",
    "L'eventuale perdita di integrità dei dati del servizio TIC può comportare interruzione di progetti di data analytics/AI, perdita di affidabilità nei confronti di partner e investitori?",
]

# ---------------------------------------------------------------------------
# Mapping livello rischio
# ---------------------------------------------------------------------------
LIVELLI_RISCHIO = {
    0: ("BASSO", "#4caf50"),
    1: ("BASSO", "#4caf50"),
    2: ("MEDIO BASSO", "#ffeb3b"),
    3: ("MEDIO ALTO", "#ff9800"),
    4: ("ALTO", "#f44336"),
    5: ("ALTO", "#f44336"),
}


def calcola_livello(punteggio: int) -> tuple[str, str]:
    return LIVELLI_RISCHIO.get(punteggio, ("ALTO", "#f44336"))


# ---------------------------------------------------------------------------
# Funzioni di rendering domande
# ---------------------------------------------------------------------------

def render_domande_si_no(titolo_sezione: str, domande: list[str], key_prefix: str) -> dict:
    """Renderizza una sezione di domande Si/No e restituisce le risposte."""
    st.subheader(titolo_sezione)
    risposte = {}
    for i, domanda in enumerate(domande):
        key = f"{key_prefix}_{i}"
        risposte[domanda] = st.radio(
            domanda,
            options=["No", "Sì"],
            index=0,
            key=key,
            horizontal=True,
        )
    return risposte


def punteggio_si_no(risposte: dict) -> int:
    return sum(1 for v in risposte.values() if v == "Sì")


# ---------------------------------------------------------------------------
# Generazione PDF
# ---------------------------------------------------------------------------

def _san(text: str) -> str:
    """Sostituisce caratteri Unicode non supportati da Helvetica (Latin-1)."""
    return (
        str(text)
        .replace("\u2019", "'").replace("\u2018", "'")
        .replace("\u201c", '"').replace("\u201d", '"')
        .replace("\u2013", "-").replace("\u2014", "-")
        .replace("\u2026", "...").replace("\u2192", "->")
        .replace("\u2190", "<-").replace("\u00b7", "-")
    )


class PDFReport(FPDF):
    def header(self):
        y_start = self.y
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, "Questionario Rischio ICT", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.cell(0, 6, f"Data: {date.today().strftime('%d/%m/%Y')}", new_x="LMARGIN", new_y="NEXT")
        # Logo in alto a destra (posizionamento assoluto, non sposta il cursore)
        logo_bytes = _logo_bytes()
        if logo_bytes:
            self.image(io.BytesIO(logo_bytes), x=self.w - 60, y=y_start, w=50)
        self.set_x(self.l_margin)
        self.ln(6)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.cell(0, 10, f"Pagina {self.page_no()}", align="C")


def genera_pdf(info: dict, categorie: dict[str, dict], punteggi: dict[str, int], livello: str) -> bytes:
    pdf = PDFReport()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    w = pdf.w - pdf.l_margin - pdf.r_margin

    # Info servizio
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(w, 8, "Informazioni Servizio")
    pdf.set_font("Helvetica", "", 10)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 6, _san(f"Nome: {info['nome']}"))
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 6, _san(f"Descrizione: {info['descrizione']}"))
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 6, _san(f"Gestito autonomamente: {'Vero' if info['d1'] else 'Falso'}"))
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 6, _san(f"Supporto funzione essenziale: {'Vero' if info['d2'] else 'Falso'}"))
    pdf.ln(4)

    # Domande e risposte per categoria
    for cat_nome, risposte in categorie.items():
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(w, 8, _san(cat_nome))
        pdf.set_font("Helvetica", "", 9)
        for domanda, risposta in risposte.items():
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(w, 5, _san(f"  D: {domanda}"))
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(w, 5, _san(f"  R: {risposta}"))
            pdf.set_font("Helvetica", "", 9)
            pdf.ln(1)
        punteggio = punteggi.get(cat_nome, 0)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(w, 6, _san(f"  Punteggio categoria: {punteggio}"))
        pdf.ln(3)

    # Risultato finale
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 13)
    max_punteggio = max(punteggi.values()) if punteggi else 0
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(w, 10, _san(f"LIVELLO DI RISCHIO: {livello} (punteggio max: {max_punteggio})"))

    return bytes(pdf.output())


# ---------------------------------------------------------------------------
# Generazione Excel
# ---------------------------------------------------------------------------

def genera_excel(info: dict, categorie: dict[str, dict], punteggi: dict[str, int], livello: str) -> bytes:
    wb = Workbook()

    # Foglio Info Servizio
    ws_info = wb.active
    ws_info.title = "Info Servizio"

    # Logo in alto
    if os.path.exists(LOGO_PATH):
        img = XlImage(LOGO_PATH)
        img.width = 250
        img.height = 63
        ws_info.add_image(img, "A1")
    # Lascia spazio sotto il logo
    ws_info.append([])
    ws_info.append([])
    ws_info.append([])
    ws_info.append([])
    ws_info.append(["Campo", "Valore"])
    row_header = ws_info.max_row
    ws_info[f"A{row_header}"].font = Font(bold=True)
    ws_info[f"B{row_header}"].font = Font(bold=True)
    ws_info.append(["Nome servizio", info["nome"]])
    ws_info.append(["Descrizione", info["descrizione"]])
    ws_info.append(["Gestito autonomamente", "Vero" if info["d1"] else "Falso"])
    ws_info.append(["Supporto funzione essenziale", "Vero" if info["d2"] else "Falso"])
    ws_info.column_dimensions["A"].width = 30
    ws_info.column_dimensions["B"].width = 60

    # Foglio Risposte
    ws_risp = wb.create_sheet("Risposte")
    ws_risp.append(["Categoria", "Domanda", "Risposta"])
    for col in ["A", "B", "C"]:
        ws_risp[f"{col}1"].font = Font(bold=True)
    for cat_nome, risposte in categorie.items():
        for domanda, risposta in risposte.items():
            ws_risp.append([cat_nome, domanda, risposta])
    ws_risp.column_dimensions["A"].width = 30
    ws_risp.column_dimensions["B"].width = 80
    ws_risp.column_dimensions["C"].width = 15

    # Foglio Risultato
    ws_res = wb.create_sheet("Risultato")
    ws_res.append(["Categoria", "Punteggio"])
    ws_res["A1"].font = Font(bold=True)
    ws_res["B1"].font = Font(bold=True)
    for cat_nome, punt in punteggi.items():
        ws_res.append([cat_nome, punt])
    max_punteggio = max(punteggi.values()) if punteggi else 0
    ws_res.append([])
    ws_res.append(["Punteggio massimo", max_punteggio])
    ws_res.append(["Livello di rischio", livello])
    # Evidenzia risultato
    last_row = ws_res.max_row
    ws_res[f"B{last_row}"].font = Font(bold=True, size=14)
    ws_res.column_dimensions["A"].width = 30
    ws_res.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# APP PRINCIPALE
# ===========================================================================

st.title("Questionario Rischio ICT")
st.markdown("Valutazione del livello di rischio per i servizi TIC bancari.")
st.divider()

# ---------------------------------------------------------------------------
# Step 1: Informazioni Generali
# ---------------------------------------------------------------------------
st.header("1. Informazioni Generali sul Servizio")

nome_servizio = st.text_input("Nome servizio / prodotto")
descrizione_servizio = st.text_area("Descrizione del servizio")

col1, col2 = st.columns(2)
with col1:
    d1 = st.radio(
        "Il servizio TIC è gestito autonomamente dalla Società?",
        options=["Vero", "Falso"],
        index=0,
        key="d1",
        horizontal=True,
    )
with col2:
    d2 = st.radio(
        "Il servizio TIC è a supporto di una funzione essenziale o importante?",
        options=["Vero", "Falso"],
        index=0,
        key="d2",
        horizontal=True,
    )

d1_bool = d1 == "Vero"
d2_bool = d2 == "Vero"

st.divider()

# ---------------------------------------------------------------------------
# Step 2: Domande condizionali
# ---------------------------------------------------------------------------
st.header("2. Valutazione Rischi")

categorie_risposte: dict[str, dict] = {}
punteggi_categorie: dict[str, int] = {}

if d1_bool and d2_bool:
    # Scenario 1: Autonomo + Essenziale -> 4 categorie
    st.info("Scenario: servizio gestito autonomamente a supporto di funzione essenziale/importante.")

    risp_disp = render_domande_si_no(
        "Rischio di Disponibilità e Continuità", DOMANDE_DISPONIBILITA, "disp"
    )
    categorie_risposte["Rischio di Disponibilità e Continuità"] = risp_disp
    punteggi_categorie["Rischio di Disponibilità e Continuità"] = punteggio_si_no(risp_disp)

    risp_sic = render_domande_si_no(
        "Rischi di Sicurezza ICT", DOMANDE_SICUREZZA, "sic"
    )
    categorie_risposte["Rischi di Sicurezza ICT"] = risp_sic
    punteggi_categorie["Rischi di Sicurezza ICT"] = punteggio_si_no(risp_sic)

    risp_cam = render_domande_si_no(
        "Rischi relativi ai Cambiamenti ICT", DOMANDE_CAMBIAMENTI, "cam"
    )
    categorie_risposte["Rischi relativi ai Cambiamenti ICT"] = risp_cam
    punteggi_categorie["Rischi relativi ai Cambiamenti ICT"] = punteggio_si_no(risp_cam)

    risp_int = render_domande_si_no(
        "Rischi di Integrità dei Dati ICT", DOMANDE_INTEGRITA, "integ"
    )
    categorie_risposte["Rischi di Integrità dei Dati ICT"] = risp_int
    punteggi_categorie["Rischi di Integrità dei Dati ICT"] = punteggio_si_no(risp_int)

elif d1_bool and not d2_bool:
    # Scenario 2: Autonomo + Non Essenziale -> 2 categorie
    st.info("Scenario: servizio gestito autonomamente, NON a supporto di funzione essenziale.")

    risp_sic = render_domande_si_no(
        "Rischi di Sicurezza ICT", DOMANDE_SICUREZZA, "sic2"
    )
    categorie_risposte["Rischi di Sicurezza ICT"] = risp_sic
    punteggi_categorie["Rischi di Sicurezza ICT"] = punteggio_si_no(risp_sic)

    risp_int = render_domande_si_no(
        "Rischi di Integrità dei Dati ICT", DOMANDE_INTEGRITA, "integ2"
    )
    categorie_risposte["Rischi di Integrità dei Dati ICT"] = risp_int
    punteggi_categorie["Rischi di Integrità dei Dati ICT"] = punteggio_si_no(risp_int)

elif not d1_bool and d2_bool:
    # Scenario 3: Non Autonomo + Essenziale -> Esternalizzazione (esteso)
    st.info("Scenario: servizio esternalizzato a supporto di funzione essenziale/importante.")

    st.subheader("Rischio di Esternalizzazione")

    ragione_sociale = st.text_input(
        "Ragione sociale del Fornitore terzo di servizi TIC",
        key="rag_soc_3",
    )

    dipendenza = st.selectbox(
        "Livello di dipendenza del servizio TIC con l'operatività totale della Banca",
        options=["Dipendenza irrilevante", "Dipendenza media", "Dipendenza totale"],
        key="dip_3",
    )
    mappa_dip = {
        "Dipendenza irrilevante": 0,
        "Dipendenza media": 1,
        "Dipendenza totale": 2,
    }

    dom_ext_1 = st.radio(
        "Il fornitore di servizi ha subito delle gravi interruzioni di servizio nel corso degli ultimi anni, che hanno causato impatti negativi sul cliente (in termini operativi, reputazionali, strategici o economici)?",
        options=["No", "Sì"],
        index=0,
        key="ext3_1",
        horizontal=True,
    )
    dom_ext_2 = st.radio(
        "Il servizio tratta dati finanziari della clientela / dipendenti della Banca di proprietà della Banca (estratti conto, informazioni personali, dati sensibili ecc..)?",
        options=["No", "Sì"],
        index=0,
        key="ext3_2",
        horizontal=True,
    )
    dom_ext_3 = st.radio(
        "L'eventuale inefficienza dell'organizzazione del Fornitore dell'esternalizzazione può comportare un impatto (a livello economico, operativo, strategico) per la Banca?",
        options=["No", "Sì"],
        index=0,
        key="ext3_3",
        horizontal=True,
    )

    punt_ext = (
        mappa_dip[dipendenza]
        + (1 if dom_ext_1 == "Sì" else 0)
        + (1 if dom_ext_2 == "Sì" else 0)
        + (1 if dom_ext_3 == "Sì" else 0)
    )

    categorie_risposte["Rischio di Esternalizzazione"] = {
        "Ragione sociale fornitore": ragione_sociale,
        "Livello di dipendenza": dipendenza,
        "Gravi interruzioni del fornitore": dom_ext_1,
        "Trattamento dati finanziari": dom_ext_2,
        "Impatto inefficienza fornitore": dom_ext_3,
    }
    punteggi_categorie["Rischio di Esternalizzazione"] = punt_ext

else:
    # Scenario 4: Non Autonomo + Non Essenziale -> Esternalizzazione (ridotto)
    st.info("Scenario: servizio esternalizzato, NON a supporto di funzione essenziale.")

    st.subheader("Rischio di Esternalizzazione")

    ragione_sociale = st.text_input(
        "Ragione sociale del Fornitore terzo di servizi TIC",
        key="rag_soc_4",
    )

    dom_ext_1 = st.radio(
        "Il fornitore di servizi ha subito delle gravi interruzioni di servizio nel corso degli ultimi anni, che hanno causato impatti negativi sul cliente (in termini operativi, reputazionali, strategici o economici)?",
        options=["No", "Sì"],
        index=0,
        key="ext4_1",
        horizontal=True,
    )
    dom_ext_2 = st.radio(
        "Il servizio tratta dati finanziari della clientela / dipendenti della Banca di proprietà della Banca (estratti conto, informazioni personali, dati sensibili ecc..)?",
        options=["No", "Sì"],
        index=0,
        key="ext4_2",
        horizontal=True,
    )
    dom_ext_3 = st.radio(
        "L'eventuale inefficienza dell'organizzazione del Fornitore dell'esternalizzazione può comportare un impatto (a livello economico, operativo, strategico) per la Banca?",
        options=["No", "Sì"],
        index=0,
        key="ext4_3",
        horizontal=True,
    )

    punt_ext = (
        (1 if dom_ext_1 == "Sì" else 0)
        + (1 if dom_ext_2 == "Sì" else 0)
        + (1 if dom_ext_3 == "Sì" else 0)
    )

    categorie_risposte["Rischio di Esternalizzazione"] = {
        "Ragione sociale fornitore": ragione_sociale,
        "Gravi interruzioni del fornitore": dom_ext_1,
        "Trattamento dati finanziari": dom_ext_2,
        "Impatto inefficienza fornitore": dom_ext_3,
    }
    punteggi_categorie["Rischio di Esternalizzazione"] = punt_ext

st.divider()

# ---------------------------------------------------------------------------
# Step 3: Risultato
# ---------------------------------------------------------------------------
st.header("3. Risultato")

if st.button("Calcola Rischio", type="primary", use_container_width=True):
    if not nome_servizio.strip():
        st.warning("Inserire il nome del servizio prima di procedere.")
    elif not punteggi_categorie:
        st.warning("Nessuna categoria di rischio disponibile.")
    else:
        max_punteggio = max(punteggi_categorie.values())
        livello, colore = calcola_livello(max_punteggio)

        export_id = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Salva in session_state per export
        st.session_state["risultato"] = {
            "info": {
                "nome": nome_servizio,
                "descrizione": descrizione_servizio,
                "d1": d1_bool,
                "d2": d2_bool,
            },
            "categorie": categorie_risposte,
            "punteggi": punteggi_categorie,
            "max_punteggio": max_punteggio,
            "livello": livello,
            "colore": colore,
            "export_id": export_id,
        }

        # Salva su Supabase
        salva_submission(
            tipo=1,
            nome_azienda=nome_servizio,  # mappato su nome_servizio_ict in supabase_client
            punteggio=max_punteggio,
            livello=livello,
            valore={"BASSO": 1, "MEDIO BASSO": 2, "MEDIO ALTO": 3, "ALTO": 4}.get(livello, 0),
            dati={
                "info": st.session_state["risultato"]["info"],
                "punteggi": punteggi_categorie,
                "export_id": export_id,
            },
        )

# Mostra risultato se presente
if "risultato" in st.session_state:
    res = st.session_state["risultato"]

    # Tabella riepilogativa
    st.subheader("Riepilogo Punteggi per Categoria")
    for cat, punt in res["punteggi"].items():
        is_max = punt == res["max_punteggio"]
        prefix = "**" if is_max else ""
        suffix = " ← MAX**" if is_max else ""
        st.markdown(f"- {prefix}{cat}: {punt}{suffix}")

    # Livello di rischio
    st.markdown("---")
    st.markdown(
        f'<div style="text-align:center; padding:20px; border-radius:10px; '
        f'background-color:{res["colore"]}; color:white; font-size:24px; font-weight:bold;">'
        f'LIVELLO DI RISCHIO: {res["livello"]}<br>'
        f'<span style="font-size:16px;">(Punteggio massimo: {res["max_punteggio"]})</span>'
        f"</div>",
        unsafe_allow_html=True,
    )

    st.markdown("")
    logo_time = "n/d"
    if os.path.exists(LOGO_PATH):
        logo_time = datetime.fromtimestamp(os.path.getmtime(LOGO_PATH)).strftime("%d/%m/%Y %H:%M:%S")
    st.caption(f"Logo attivo: {os.path.basename(LOGO_PATH)} — aggiornato: {logo_time}")

    # Export
    export_id = res.get("export_id", datetime.now().strftime("%Y%m%d_%H%M%S"))
    export_base_name = f"rischio_ict_{nome_servizio.replace(' ', '_')}_{export_id}"
    col_pdf, col_xlsx = st.columns(2)
    with col_pdf:
        pdf_bytes = genera_pdf(
            res["info"], res["categorie"], res["punteggi"], res["livello"]
        )
        st.download_button(
            label="Scarica PDF",
            data=pdf_bytes,
            file_name=f"{export_base_name}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    with col_xlsx:
        xlsx_bytes = genera_excel(
            res["info"], res["categorie"], res["punteggi"], res["livello"]
        )
        st.download_button(
            label="Scarica Excel",
            data=xlsx_bytes,
            file_name=f"{export_base_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
